import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import dropbox
import os

common = {'JAN': 'A60',
          'FEB': 'A60',
          'MAR': 'A60',
          'MAY': 'A60',
          'JUN': 'A60',
          'AUG': 'A60',
          'SEP': 'A60',
          'OCT': 'A60',
          'NOV': 'A60'}

all_months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
choose_month = {'W01': 'JAN', 'W02': 'JAN', 'W03': 'JAN', 'W04': 'JAN', 'W05': 'JAN',
                'W06': 'FEB', 'W07': 'FEB', 'W08': 'FEB', 'W09': 'FEB',
                'W10': 'MAR', 'W11': 'MAR', 'W12': 'MAR', 'W13': 'MAR',
                'W14': 'APR', 'W15': 'APR', 'W16': 'APR', 'W17': 'APR', 'W18': 'APR',
                'W19': 'MAY', 'W20': 'MAY', 'W21': 'MAY', 'W22': 'MAY',
                'W23': 'JUN', 'W24': 'JUN', 'W25': 'JUN', 'W26': 'JUN',
                'W27': 'JUL', 'W28': 'JUL', 'W29': 'JUL', 'W30': 'JUL', 'W31': 'JUL',
                'W32': 'AUG', 'W33': 'AUG', 'W34': 'AUG', 'W35': 'AUG',
                'W36': 'SEP', 'W37': 'SEP', 'W38': 'SEP', 'W39': 'SEP',
                'W40': 'OCT', 'W41': 'OCT', 'W42': 'OCT', 'W43': 'OCT', 'W44': 'OCT',
                'W45': 'NOV', 'W46': 'NOV', 'W47': 'NOV', 'W48': 'NOV',
                'W49': 'DEC', 'W50': 'DEC', 'W51': 'DEC', 'W52': 'DEC'}

week_location = {'W01': 8,
                 'W02': 21,
                 'W03': 34,
                 'W04': 47,
                 'W05': 60,
                 'W06': 21,
                 'W07': 34,
                 'W08': 47,
                 'W09': 60,
                 'W10': 21,
                 'W11': 34,
                 'W12': 47,
                 'W13': 60,
                 'W14': 21,
                 'W15': 34,
                 'W16': 47,
                 'W17': 60,
                 'W18': 73,
                 'W19': 21,
                 'W20': 34,
                 'W21': 47,
                 'W22': 60,
                 'W23': 21,
                 'W24': 34,
                 'W25': 47,
                 'W26': 60,
                 'W27': 21,
                 'W28': 34,
                 'W29': 47,
                 'W30': 60,
                 'W31': 73,
                 'W32': 21,
                 'W33': 34,
                 'W34': 47,
                 'W35': 60,
                 'W36': 21,
                 'W37': 34,
                 'W38': 47,
                 'W39': 60,
                 'W40': 8,
                 'W41': 21,
                 'W42': 34,
                 'W43': 47,
                 'W44': 60,
                 'W45': 21,
                 'W46': 34,
                 'W47': 47,
                 'W48': 60,
                 'W49': 21,
                 'W50': 34,
                 'W51': 47,
                 'W52': 60}


day_coordinate = {
    # this dictionary is for source sheet
    'monday': ['3', '11'],
    'tuesday': ['12', '20'],
    'wednesday': ['21', '29'],
    'thursday': ['30', '38'],
    'friday': ['39', '47'],
    'saturday': ['48', '56'],
    'sunday': ['57', '65']}

day_coordinate2 = {
    # this dictionary is for save sheet. Normal hours offset from name cell
    'monday': 2,
    'tuesday': 4,
    'wednesday': 6,
    'thursday': 8,
    'friday': 10,
    'saturday': 12,
    'sunday': 14}


def main():
    main_menu()
    user_id = input("Wybierz cyfre aby wybrac uzytkownika\n") # taking an input from user. Integer that is used for menu navigation
    input_which_week = input("Podaj ktory tydzien chcesz uzupelnic #format typu 'W01': \n") # which week to update in "W01" format
    month = choose_month[input_which_week]
    sciezki = dropbox_connect(user_id, input_which_week) #function that connects to dropbox, downloads and saves them as copy and returns paths to source and save sheet
    # sciezki[0] - source sheet
    # sciezki[1] - save sheet
    source_sheet = openpyxl.load_workbook(sciezki[0], keep_vba=True) # create openpyxl source_sheet object
    save_sheet = openpyxl.load_workbook(sciezki[1]) # create openpyxl save_sheet object
    save_which_month = save_sheet[month] # "open" sheet by correct month
    clear_cells(input_which_week, save_which_month) # clear cells in chosen week to avoid doubling up the values
    source_which_week = source_sheet[input_which_week] # "open" sheet by correct week
    #save_which_month = save_sheet[month]
    for key in day_coordinate:
        source_data = get_source_dictionary(key, day_coordinate, source_which_week)
        save_data = get_save_dictionary(key, day_coordinate2, save_which_month, input_which_week)
        #print(save_data)
        save_test(source_data, save_data, input_which_week)
        print_to_excel(save_data, key, day_coordinate2, sciezki, save_which_month, save_sheet)
    for item in all_months:
        if item in common:
            style_days_common(item, sciezki, save_sheet)
        else:
            style_days_longer(item, sciezki, save_sheet)
    dropbox_upload(user_id, sciezki[1], input_which_week)

def clear_cells(input_which_week, save_which_month):
    ws = save_which_month
    start = 'A' + str(week_location[input_which_week])
    stop = 'P' + str(week_location[input_which_week]+4)
    #print(start, stop)
    for row in ws[start:stop]:
        for cell in row:
            cell.value = None
    #for row in ws[]

def get_source_dictionary(key, day_coordinate, source_which_week):
    """
    This function returns a dictionary full of projects in source sheet.
    key = coordinate
    [0] = proj_nro
    [1] = normal_hours
    [2] = overtime_hours
    :return: dictionary
    """
    day_start = int(day_coordinate[key][0])  #
    day_stop = int(day_coordinate[key][1])
    dictionary = {}
    for col in source_which_week.iter_cols(min_col=3, max_col=3, min_row=day_start,
                                           max_row=day_stop):
        for cell in col:
            key = cell.coordinate
            litera = cell.coordinate[:1]
            liczba = cell.coordinate[1:]
            customer = cell.value
            proj_nro = chr(ord(litera) + 1) + liczba
            normal_hours = chr(ord(litera) + 3) + liczba
            overtime_hours = chr(ord(litera) + 4) + liczba
            return_list = [customer, source_which_week[proj_nro].value, source_which_week[normal_hours].value,
                           source_which_week[overtime_hours].value]
            dictionary[key] = return_list
    return dictionary


def get_save_dictionary(key2, day_coordinate2, save_which_month, input_which_week):
    """
    This function returns a dictionary full of projects in save sheet.
    key = coordinate
    [0] = proj_nro
    [1] = normal_hours
    [2] = overtime_hours
    :return: dictionary
    """
    dictionary = {}
    for col in save_which_month.iter_cols(min_col=1, max_col=1, min_row=week_location[input_which_week],
                                          max_row=week_location[input_which_week] + 4):
        for cell in col:
            key = cell.coordinate
            litera = cell.coordinate[:1]
            liczba = cell.coordinate[1:]
            customer = cell.value
            proj_nro = chr(ord(litera) + 1) + liczba
            normal_hours = chr(ord(litera) + day_coordinate2[key2]) + liczba
            overtime_hours = chr(ord(litera) + day_coordinate2[key2] + 1) + liczba
            return_list = [customer, save_which_month[proj_nro].value, save_which_month[normal_hours].value,
                           save_which_month[overtime_hours].value]
            dictionary[key] = return_list
    return dictionary


def find_last_index(save_data, input_which_week):
    for j in range(week_location[input_which_week],
                   week_location[input_which_week] + 5):  # 8 13 before changes to week_location
        index = 'A' + str(j)
        if save_data[index][0] == None:
            # print(index + ' is empty')
            return index


def save_test(source_data, save_data, input_which_week):
    for key_source in list(source_data.keys()):
        i = 0
        for key_save in list(
                save_data.keys()):  # this list magic thing "avoids dictionary changed size during iteration" error
            if save_data[key_save][0] == source_data[key_source][0]:
                if source_data[key_source][0] == 'Specific_client_name_it_starts_on_V - CA/MWA':
                    if save_data[key_save][2] != None:
                        normal_hours = float(save_data[key_save][2])
                    else:
                        normal_hours = 0
                    if source_data[key_source][2] != None:
                        normal_hours += float(source_data[key_source][2])
                    if save_data[key_save][3] != None:
                        overtime_hours = float(save_data[key_save][3])
                    else:
                        overtime_hours = 0
                    if source_data[key_source][3] != None:
                        overtime_hours += float(source_data[key_source][3])
                    volvo_updated_list = ['Specific_client_name_it_starts_on_V - CA/MWA', None, normal_hours, overtime_hours]
                    save_data[key_save] = volvo_updated_list
                    i += 1  # this helps. Without it it keeps on adding few volvos
                else:
                    if save_data[key_save][1] == source_data[key_source][1]:
                        if save_data[key_save][2] != None:
                            normal_hours = float(save_data[key_save][2])
                        else:
                            normal_hours = 0
                        if source_data[key_source][2] != None:
                            normal_hours += float(source_data[key_source][2])
                        if save_data[key_save][3] != None:
                            overtime_hours = float(save_data[key_save][3])
                        else:
                            overtime_hours = 0
                        if source_data[key_source][3] != None:
                            overtime_hours += float(source_data[key_source][3])
                        updated_list = [save_data[key_save][0], save_data[key_save][1], normal_hours, overtime_hours]
                        save_data[key_save] = updated_list

                        i += 1
                        break

        if i == 0:
            x = find_last_index(save_data, input_which_week)
            if x != None:
                save_data[x] = source_data[key_source]
                # print(x, key_source)
    # print(save_data)
    return save_data


def print_to_excel(save_data, key2, day_coordinate2, sciezki, save_which_month, save_sheet):
    for key in save_data:
        name_index = key
        proj_nro_index = 'B' + key[1:]
        normal_hours_index = chr(ord('A') + day_coordinate2[key2]) + key[1:]
        overtime_hours_index = chr(ord('A') + day_coordinate2[key2] + 1) + key[1:]
        name = save_data[key][0]
        proj_nro = save_data[key][1]
        normal_hours = save_data[key][2]
        overtime_hours = save_data[key][3]
        save_which_month[name_index] = name
        save_which_month[proj_nro_index] = proj_nro
        save_which_month[normal_hours_index] = normal_hours
        save_which_month[overtime_hours_index] = overtime_hours
    save_sheet.save(sciezki[1])
    save_sheet.close()
    # print(save_data)


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def style_week_days(i, interval, ws, border, fill, font):
    i -= 1
    for j in range(1, 14, 2):
        dzien_tygodnia = chr(66 + j) + str(5 + i * interval) + ':' + chr(67 + j) + str(5 + i * interval)
        style_range(ws, dzien_tygodnia, border=border, fill=fill, font=font)
        dzien_tygodnia2 = chr(66 + j) + str(6 + i * interval) + ':' + chr(67 + j) + str(6 + i * interval)
        style_range(ws, dzien_tygodnia2, border=border, fill=fill, font=font)


def style_days_common(which_month, sciezki, save_sheet):
    wb = save_sheet
    ws = save_sheet[which_month]
    thin = Side(border_style="thin", color="000000")
    none = Side(border_style="none", color="000000")
    medium = Side(border_style="medium", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    border2 = Border(top=none, left=medium, right=thin, bottom=thin)
    border3 = Border(top=thin, left=thin, right=medium, bottom=thin)
    border4 = Border(top=thin, left=medium, right=medium, bottom=thin)
    border5 = Border(top=none, left=medium, right=none, bottom=none)
    border6 = Border(top=medium, left=thin, right=thin, bottom=thin)
    border7 = Border(top=none, left=thin, right=medium, bottom=medium)
    border8 = Border(top=thin, left=none, right=medium, bottom=medium)
    border9 = Border(top=none, left=none, right=none, bottom=medium)
    border12 = Border(top=medium, left=medium, right=medium, bottom=medium)
    border13 = Border(top=medium, left=thin, right=medium, bottom=thin)
    border14 = Border(top=medium, left=medium, right=thin, bottom=thin)
    border15 = Border(top=none, left = none, right = none, bottom = none)
    fill = PatternFill("solid", fgColor="C0C0C0")
    fill_green = PatternFill("solid", fgColor="008080")
    fill_orange = PatternFill("solid", fgColor="FFCC99")
    fill_blue = PatternFill("solid", fgColor="ccffff")
    font = Font(b=False, color="000000")
    font_bold = Font(b=True, color="000000")
    font_gray = Font(b=False, color="808080")
    al = Alignment(horizontal="center", vertical="center")
    al2 = Alignment(horizontal="center", vertical="bottom")
    al_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al_text_wrap = Alignment(wrap_text=True)
    interval = 13
    for i in range(1, 6):
        if i < 6:
            duza_ramka = 'S' + str(i * interval) + ':' + 'V' + str(i * interval)
            nominal = 'A' + str(i * interval) + ':' + 'R' + str(i * interval)
            sobota = 'M' + str(8 + (i - 1) * interval) + ':' + 'N' + str(13 + (i - 1) * interval)
            niedziela = 'O' + str(8 + (i - 1) * interval) + ':' + 'P' + str(13 + (i - 1) * interval)
            total = 'Q' + str(5 + (i - 1) * interval) + ':' + 'R' + str(6 + (i - 1) * interval)
            customer = 'A' + str(5 + (i - 1) * interval) + ':' + 'A' + str(7 + (i - 1) * interval)
            proj_nro = 'B' + str(5 + (i - 1) * interval) + ':' + 'B' + str(7 + (i - 1) * interval)
            notes = 'C' + str(3 + i * interval) + ':' + 'R' + str(3 + i * interval)
            approved = 'A' + str(2 + i * interval) + ':' + 'B' + str(3 + i * interval)
            plus_minus = 'M' + str(2 + i * interval) + ':' + 'N' + str(2 + i * interval)
            style_range(ws, duza_ramka, border=border4, fill=fill, font=font)
            style_range(ws, nominal, border=border, fill=fill, font=font)
            style_range(ws, sobota, border=border, fill=fill, font=font)
            style_range(ws, niedziela, border=border, fill=fill, font=font)
            style_range(ws, total, border=border3, fill=fill, font=font, alignment=al)
            style_range(ws, customer, border=border, fill=fill, font=font_bold, alignment=al2)
            style_range(ws, proj_nro, border=border, fill=fill, font=font_bold, alignment=al2)
            style_range(ws, notes, border=border7, fill=fill, font=font)
            style_range(ws, approved, border=border8, fill=fill, font=font)
            style_range(ws, plus_minus, border=border, fill=fill, font=font)
        style_week_days(i, interval, ws, border, fill, font)
    style_range(ws, 'T5:T7', border=border, fill=fill, font=font, alignment=al)
    style_range(ws, 'U5:U7', border=border, fill=fill, font=font, alignment=al)
    style_range(ws, 'S2:S4', border=border14, fill=fill_green, font=font, alignment=al_wrap)
    style_range(ws, 'T2:T4', border=border6, fill=fill_green, font=font, alignment=al_wrap)
    style_range(ws, 'U2:U4', border=border, fill=fill_green, font=font, alignment=al)
    style_range(ws, 'V2:V4', border=border13, fill=fill_orange, font=font, alignment=al)
    style_range(ws, 'V5:V7', border=border, fill=fill, font=font, alignment=al)
    style_range(ws, 'W5:W80', border=border5, font=font)
    style_range(ws, 'A2:B4', border=border6, fill=fill, font=font)
    style_range(ws, 'A1:V1', border=border9, font=font)
    style_range(ws, 'S72:T72', border=border, fill=fill_green, font=font)
    style_range(ws, 'S73:T73', border=border, fill=fill_green, font=font)
    style_range(ws, 'S74:T74', border=border, fill=fill_green, font=font)
    style_range(ws, 'S75:T75', border=border, fill=fill_green, font=font)
    style_range(ws, 'S77:T77', border=border, fill=fill_orange, font=font)
    style_range(ws, 'S78:T78', border=border, fill=fill_orange, font=font)
    style_range(ws, 'S79:T79', border=border, fill=fill_orange, font=font)
    style_range(ws, 'S80:T80', border=border, fill=fill_orange, font=font)
    style_range(ws, 'C5:P7', border=border, fill=fill, font=font)
    style_range(ws, 'S5:S7', border=border2, fill=fill, font=font, alignment=al)
    style_range(ws, 'H70:I70', border=border12, fill=fill_blue, font=font_bold)
    style_range(ws, 'H71:I71', border=border12, fill=fill, font=font)
    style_range(ws, 'J71:P71', border=border12, fill=fill, font=font)
    style_range(ws, 'Q71:R71', border=border12, fill=fill, font=font)
    style_range(ws, 'U72:V72', border=border, font=font)
    style_range(ws, 'U73:V73', border=border, font=font)
    style_range(ws, 'U74:V74', border=border, font=font)
    style_range(ws, 'U75:V75', border=border, font=font)
    style_range(ws, 'U77:V77', border=border, font=font)
    style_range(ws, 'U78:V78', border=border, font=font)
    style_range(ws, 'U79:V79', border=border, font=font)
    style_range(ws, 'U80:V80', border=border, font=font)
    style_range(ws, 'S71:V71', border=border12, fill=fill_green, font=font)
    style_range(ws, 'C70:F70', border=border15, font=font_gray)
    if which_month == 'OCT':
        style_range(ws, 'S17:V20', border=border4, fill=fill, font=font)
        style_range(ws, 'S56:V59', border=border4, fill=fill, font=font)
    for k in range(79, 87):
        x = 'C' + str(k) + ':' + 'N' + str(k)
        style_range(ws, x, border=border, font=font_gray)
    for l in range(88, 105):
        x = 'C' + str(l) + ':' + 'N' + str(l)
        style_range(ws, x, border=border, font=font_gray)
    for m in range(79, 87):
        x = 'O' + str(m) + ':' + 'P' + str(m)
        style_range(ws, x, border=border, font=font_gray)
    for n in range(88, 105):
        x = 'O' + str(n) + ':' + 'P' + str(n)
        style_range(ws, x, border=border, font=font_gray)
    ws['C4'] = 'test'
    wb.save(sciezki[1])


def style_days_longer(which_month, sciezki, save_sheet):
    wb = save_sheet
    ws = save_sheet[which_month]
    thin = Side(border_style="thin", color="000000")
    none = Side(border_style="none", color="000000")
    medium = Side(border_style="medium", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    border2 = Border(top=none, left=medium, right=thin, bottom=thin)
    border3 = Border(top=thin, left=thin, right=medium, bottom=thin)
    border4 = Border(top=thin, left=medium, right=medium, bottom=thin)
    border5 = Border(top=none, left=medium, right=none, bottom=none)
    border6 = Border(top=medium, left=thin, right=thin, bottom=thin)
    border7 = Border(top=none, left=thin, right=medium, bottom=medium)
    border8 = Border(top=thin, left=none, right=medium, bottom=medium)
    border9 = Border(top=none, left=none, right=none, bottom=medium)
    border12 = Border(top=medium, left=medium, right=medium, bottom=medium)
    border13 = Border(top=medium, left=thin, right=medium, bottom=thin)
    border14 = Border(top=medium, left=medium, right=thin, bottom=thin)
    border15 = Border(top=none, left=none, right=none, bottom=none)
    fill = PatternFill("solid", fgColor="C0C0C0")
    fill_green = PatternFill("solid", fgColor="008080")
    fill_orange = PatternFill("solid", fgColor="FFCC99")
    fill_blue = PatternFill("solid", fgColor="ccffff")
    font = Font(b=False, color="000000")
    font_bold = Font(b=True, color="000000")
    font_gray = Font(b=False, color="808080")
    al = Alignment(horizontal="center", vertical="center")
    al2 = Alignment(horizontal="center", vertical="bottom")
    al_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    interval = 13
    for i in range(1, 7):
        if i < 7:
            duza_ramka = 'S' + str(i * interval) + ':' + 'V' + str(i * interval)
            nominal = 'A' + str(i * interval) + ':' + 'R' + str(i * interval)
            sobota = 'M' + str(8 + (i - 1) * interval) + ':' + 'N' + str(13 + (i - 1) * interval)
            niedziela = 'O' + str(8 + (i - 1) * interval) + ':' + 'P' + str(13 + (i - 1) * interval)
            total = 'Q' + str(5 + (i - 1) * interval) + ':' + 'R' + str(6 + (i - 1) * interval)
            customer = 'A' + str(5 + (i - 1) * interval) + ':' + 'A' + str(7 + (i - 1) * interval)
            proj_nro = 'B' + str(5 + (i - 1) * interval) + ':' + 'B' + str(7 + (i - 1) * interval)
            notes = 'C' + str(3 + i * interval) + ':' + 'R' + str(3 + i * interval)
            approved = 'A' + str(2 + i * interval) + ':' + 'B' + str(3 + i * interval)
            plus_minus = 'M' + str(2 + i * interval) + ':' + 'N' + str(2 + i * interval)
            style_range(ws, duza_ramka, border=border4, fill=fill, font=font)
            style_range(ws, nominal, border=border, fill=fill, font=font)
            style_range(ws, sobota, border=border, fill=fill, font=font)
            style_range(ws, niedziela, border=border, fill=fill, font=font)
            style_range(ws, total, border=border3, fill=fill, font=font, alignment=al)
            style_range(ws, customer, border=border, fill=fill, font=font_bold, alignment=al2)
            style_range(ws, proj_nro, border=border, fill=fill, font=font_bold, alignment=al2)
            style_range(ws, notes, border=border7, fill=fill, font=font)
            style_range(ws, approved, border=border8, fill=fill, font=font)
            style_range(ws, plus_minus, border=border, fill=fill, font=font)
        style_week_days(i, interval, ws, border, fill, font)
    style_range(ws, 'T5:T7', border=border, fill=fill, font=font, alignment=al)
    style_range(ws, 'U5:U7', border=border, fill=fill, font=font, alignment=al)
    style_range(ws, 'S2:S4', border=border14, fill=fill_green, font=font, alignment=al_wrap)
    style_range(ws, 'T2:T4', border=border6, fill=fill_green, font=font, alignment=al_wrap)
    style_range(ws, 'U2:U4', border=border, fill=fill_green, font=font, alignment=al)
    style_range(ws, 'V2:V4', border=border13, fill=fill_orange, font=font, alignment=al)
    style_range(ws, 'V5:V7', border=border, fill=fill, font=font, alignment=al)
    style_range(ws, 'W5:W80', border=border5, font=font)
    style_range(ws, 'A2:B4', border=border6, fill=fill, font=font)
    style_range(ws, 'A1:V1', border=border9, font=font)
    style_range(ws, 'S85:T85', border=border, fill=fill_green, font=font)
    style_range(ws, 'S86:T86', border=border, fill=fill_green, font=font)
    style_range(ws, 'S87:T87', border=border, fill=fill_green, font=font)
    style_range(ws, 'S88:T88', border=border, fill=fill_green, font=font)
    style_range(ws, 'S90:T90', border=border, fill=fill_orange, font=font)
    style_range(ws, 'S91:T91', border=border, fill=fill_orange, font=font)
    style_range(ws, 'S92:T92', border=border, fill=fill_orange, font=font)
    style_range(ws, 'S93:T93', border=border, fill=fill_orange, font=font)
    style_range(ws, 'C5:P7', border=border, fill=fill, font=font)
    style_range(ws, 'S5:S7', border=border2, fill=fill, font=font, alignment=al)
    style_range(ws, 'H83:I83', border=border12, fill=fill_blue, font=font_bold)
    style_range(ws, 'H84:I84', border=border12, fill=fill, font=font)
    style_range(ws, 'J84:P84', border=border12, fill=fill, font=font)
    style_range(ws, 'Q84:R84', border=border12, fill=fill, font=font)
    style_range(ws, 'U85:V85', border=border, font=font)
    style_range(ws, 'U86:V86', border=border, font=font)
    style_range(ws, 'U87:V87', border=border, font=font)
    style_range(ws, 'U88:V88', border=border, font=font)
    style_range(ws, 'U90:V90', border=border, font=font)
    style_range(ws, 'U91:V91', border=border, font=font)
    style_range(ws, 'U92:V92', border=border, font=font)
    style_range(ws, 'U93:V93', border=border, font=font)
    style_range(ws, 'S84:V84', border=border12, fill=fill_green, font=font)
    style_range(ws, 'S69:V72', border=border, fill=fill, font=font)
    style_range(ws, 'C83:F83', border=border15, font=font_gray)
    if which_month == 'OCT':
        style_range(ws, 'S17:V20', border=border4, font=font)
        style_range(ws, 'S56:V59', border=border4, font=font)
    for k in range(92, 100):
        x = 'C' + str(k) + ':' + 'N' + str(k)
        style_range(ws, x, border=border, font=font_gray)
    for l in range(101, 117):
        x = 'C' + str(l) + ':' + 'N' + str(l)
        style_range(ws, x, border=border, font=font_gray)
    for m in range(92, 100):
        x = 'O' + str(m) + ':' + 'P' + str(m)
        style_range(ws, x, border=border, font=font_gray)
    for n in range(101, 117):
        x = 'O' + str(n) + ':' + 'P' + str(n)
        style_range(ws, x, border=border, font=font_gray)
    wb.save(sciezki[1])



def main_menu():
    print('Wybierz numer aby wskazac komu podliczyc godziny:')
    global user_list
    user_list = ['1. user1','2. user2','3. user3','4. user4','5. user5','6. user6','7. user7']
    for user_name in user_list:
        print(user_name)


def dropbox_connect(user_id, input_which_week):
    indeks = int(user_id) - 1
    global path_folder
    path_folder = ['/LDA/','/TSO/','/PMO/','/MMA/','/PZO/','/PST/','/DMA/']
    global sheet_name_save
    sheet_name_save = ['user1_2018.xlsx','user2_2018.xlsx','user3_2018.xlsx','user4_2018.xlsx',
                       'user5_2018.xlsx','user6_2018.xlsx','user7_2018.xlsx']
    sheet_name_source = ['user1_LDA_2018.xlsm', 'user2_TSO_2018.xlsm', 'user3_PMO_2018.xlsm', 'user4_MMA_2018.xlsm',
                         'user5_PZO_2018.xlsm', 'user6_PST_2018.xlsm', 'user7_DMA_2018.xlsm']
    global token
    token = ['list of safety tokens here']
    dbx = dropbox.Dropbox(token[0]) #TODO zmienic tutaj po uzupelnieniu listy tokenow
    #print(dbx.users_get_current_account())
    print('Laczenie z dropboxem...')
    #print(dbx.files_list_folder(path_folder[indeks]).entries)
    for entry in dbx.files_list_folder(path_folder[indeks]).entries:
        if entry.name == sheet_name_source[indeks]:  # sprawdza czy na dropie jest poprawnie nazwany plik
            full_path_to_dropbox_source = path_folder[indeks] + sheet_name_source[indeks]
            full_path_to_local_source = 'C:\Timesheet_old\\Source\\' + str(sheet_name_source[indeks].split('.')[0]) + '_' + input_which_week + '.xlsm'
            full_path_to_dropbox_save = path_folder[indeks] + sheet_name_save[indeks]
            full_path_to_local_save = 'C:\Timesheet_old\\Save\\' + str(sheet_name_save[indeks].split('.')[0]) + '_' + input_which_week + '.xlsx'
            print('Sciaganie arkuszy z dropboxa...')
            if os.path.exists('C:\Timesheet_old\\Source\\') != True:
                os.mkdir('C:\Timesheet_old\\Source\\')
            if os.path.exists('C:\Timesheet_old\\Save\\') != True:
                os.mkdir('C:\Timesheet_old\\Save\\')
            print(full_path_to_dropbox_save)
            print(full_path_to_dropbox_source)
            dbx.files_download_to_file(full_path_to_local_source, full_path_to_dropbox_source)
            #time.sleep(5)
            dbx.files_download_to_file(full_path_to_local_save, full_path_to_dropbox_save)
        else:
            pass
    print('Tworzenie kopii zapasowej dla tygodnia {}...'.format(input_which_week))
    return full_path_to_local_source, full_path_to_local_save

def dropbox_upload(user_id, save_path, input_which_week):
    indeks = int(user_id) - 1
    dbx = dropbox.Dropbox(token[0])
    with open(save_path, 'rb') as file:
        dbx.files_upload(file.read(), path_folder[indeks] + str(sheet_name_save[indeks].split('.')[0])+'.xlsx', mode=dropbox.files.WriteMode.overwrite)
        print('Arkusz godzinowy dla uzytkownika {} zostal podliczony dla tygodnia {}'.format(user_list[indeks], input_which_week))
    file.close()


if __name__ == "__main__":
    main()
